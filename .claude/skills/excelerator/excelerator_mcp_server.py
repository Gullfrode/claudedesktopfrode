#!/usr/bin/env python3
"""
Excelerator GL MCP Server – UBW Agresso / Sigma2
Genererer Excelerator-kompatible GL-bilag (.xlsx) klar for import via Citrix.
"""

import os
import glob
import json
from datetime import datetime
from pathlib import Path
from mcp.server.fastmcp import FastMCP

mcp = FastMCP("excelerator-gl")

DESKTOP = Path.home() / "Desktop"


# ---------------------------------------------------------------------------
# Core generation logic (from generate_gl_bilag.py)
# ---------------------------------------------------------------------------

def _neste_batch_id(periode: int) -> int:
    """Finn høyeste brukte batch_id for perioden og inkrementer."""
    try:
        import openpyxl
    except ImportError:
        return int(f"{periode}00")

    høyeste = -1
    for fil in glob.glob(str(DESKTOP / "*.xlsx")):
        if os.path.basename(fil).startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(fil, data_only=True)
        except Exception:
            continue
        if "_control" in wb.sheetnames:
            ws = wb["_control"]
            for row in ws.iter_rows(values_only=True):
                if row[0] == "setdefault" and row[1] == "batch_id":
                    bid = row[2]
                    if isinstance(bid, int) and str(bid).startswith(str(periode)):
                        løpenr = int(str(bid)[6:])
                        høyeste = max(høyeste, løpenr)
    return int(f"{periode}{(høyeste + 1):02d}")


def _generer_fil(
    periode: int,
    bokføringsdato: datetime,
    linjer: list[dict],
    bilagsnavn: str | None = None,
    batch_id: int | None = None,
) -> str:
    """Generer xlsx og returner filsti."""
    import openpyxl
    from openpyxl.styles import Font

    total = sum(l["amount"] for l in linjer)
    if abs(total) > 0.01:
        raise ValueError(f"Bilaget balanserer ikke: sum = {total:.2f} (må være 0)")
    if not linjer:
        raise ValueError("Ingen posteringslinjer")
    for i, l in enumerate(linjer):
        for felt in ("account", "amount", "description"):
            if felt not in l:
                raise ValueError(f"Linje {i+1} mangler påkrevd felt: {felt}")

    if batch_id is None:
        batch_id = _neste_batch_id(periode)

    wb = openpyxl.Workbook()
    ws_ctrl = wb.active
    ws_ctrl.title = "_control"

    ws_ctrl["A1"] = "*"
    ws_ctrl["B1"] = "BOKFØRING AV BILAG FRA EXCEL"
    ws_ctrl["A4"] = "*"
    ws_ctrl["B4"] = "Global Parameters (setdefault will be used unless parameter of same name is passed in from Agresso)"
    ws_ctrl["A5"] = "*"
    ws_ctrl["B5"] = "Parameter"
    ws_ctrl["C5"] = "Value"

    params = [
        ("client",         "client",         "FIRMA KODE"),
        ("batch_id",       batch_id,         "BUNT NUMMER/ID"),
        ("period",         periode,           "Periode"),
        ("voucher_date",   bokføringsdato,    "Bilagsdato"),
        ("voucher_no",     None,              "Bilagsnummer"),
        ("voucher_type",   "GL",              "Bilagsart"),
        ("user_id",        "user_id",         "Bruker ID"),
        ("currency",       "NOK",             "Valuta"),
        ("vouch_flag",     "Y",               "Skal GL07 tildele bilags nr? (Y/N)"),
        ("variant_number", 9,                 "post back paramter for GL07 variant"),
        ("trans_type",     "GL",              "Hovedbilag = GL"),
        ("interface",      "BI",              "Forsystem"),
    ]
    for rad, (key, val, kommentar) in enumerate(params, start=6):
        ws_ctrl.cell(rad, 1, "setdefault")
        ws_ctrl.cell(rad, 2, key)
        ws_ctrl.cell(rad, 3, val)
        ws_ctrl.cell(rad, 4, kommentar)

    ws_post = wb.create_sheet("Postering til UBW")
    ws_post["C2"] = "Hovedbokstransaksjoner"

    kolonner = ["update_columns", "account", "dim_1", "dim_2", "dim_3",
                "dim_4", "dim_6", "tax_code", "amount", "cur_amount", "description"]
    for col, verdi in enumerate(kolonner, start=1):
        ws_post.cell(9, col, verdi).font = Font(bold=True)

    for rad_offset, linje in enumerate(linjer):
        rad = 10 + rad_offset
        ws_post.cell(rad, 1, "update_data")
        ws_post.cell(rad, 2, linje["account"])
        ws_post.cell(rad, 3, linje.get("dim_1"))
        ws_post.cell(rad, 4, linje.get("dim_2"))
        ws_post.cell(rad, 5, linje.get("dim_3"))
        ws_post.cell(rad, 6, linje.get("dim_4"))
        ws_post.cell(rad, 7, linje.get("dim_6"))
        ws_post.cell(rad, 8, linje.get("tax_code"))
        ws_post.cell(rad, 9, linje["amount"])
        ws_post.cell(rad, 10, linje["amount"])
        ws_post.cell(rad, 11, linje["description"])

    if bilagsnavn is None:
        år = str(periode)[:4]
        mnd = str(periode)[4:6]
        bilagsnavn = f"{år} {mnd} Bilag {batch_id}"

    filsti = DESKTOP / f"{bilagsnavn}.xlsx"
    wb.save(filsti)
    return str(filsti)


def _parse_periode(periode_str: str) -> int:
    """Konverter periode-streng til int YYYYMM."""
    s = str(periode_str).strip()
    if s.isdigit() and len(s) == 6:
        return int(s)
    måneder = {
        "januar": 1, "februar": 2, "mars": 3, "april": 4,
        "mai": 5, "juni": 6, "juli": 7, "august": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12,
    }
    s_lower = s.lower()
    for navn, nr in måneder.items():
        if navn in s_lower:
            år = datetime.today().year
            return int(f"{år}{nr:02d}")
    raise ValueError(f"Ugyldig periode: '{periode_str}'. Bruk YYYYMM eller månedsnavn.")


def _parse_dato(dato_str: str) -> datetime:
    """Konverter dato-streng til datetime."""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(dato_str.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Ugyldig dato: '{dato_str}'. Bruk DD.MM.YYYY eller YYYY-MM-DD.")


# ---------------------------------------------------------------------------
# MCP Tools
# ---------------------------------------------------------------------------

@mcp.tool()
def preview_gl_bilag(
    periode: str,
    bokføringsdato: str,
    linjer_json: str,
) -> str:
    """
    Forhåndsvis et GL-bilag uten å generere fil. Validerer at bilaget balanserer.

    Args:
        periode: Periode som YYYYMM eller månedsnavn (f.eks. "202603" eller "mars 2026")
        bokføringsdato: Bokføringsdato som DD.MM.YYYY eller YYYY-MM-DD
        linjer_json: JSON-array med posteringslinjer. Hvert objekt må ha:
            account (int), amount (float), description (str).
            Valgfritt: dim_1, dim_2, dim_3, dim_4, dim_6, tax_code.
            Eksempel: [{"account": 29300, "amount": -9789.0, "description": "Test"},
                       {"account": 29300, "amount":  9789.0, "description": "Test"}]
    """
    try:
        periode_int = _parse_periode(periode)
        dato = _parse_dato(bokføringsdato)
        linjer = json.loads(linjer_json)
    except Exception as e:
        return f"Feil i input: {e}"

    total = sum(l.get("amount", 0) for l in linjer)
    batch_id = _neste_batch_id(periode_int)

    lines_out = []
    for i, l in enumerate(linjer, 1):
        beløp = l.get("amount", 0)
        dr_kr = "Debet" if beløp > 0 else "Kredit"
        lines_out.append(
            f"  {i}. konto={l.get('account')}  {dr_kr} {abs(beløp):,.2f}  "
            f"dim1={l.get('dim_1','')}  dim2={l.get('dim_2','')}  "
            f"dim3={l.get('dim_3','')}  '{l.get('description','')}'"
        )

    status = "OK – bilaget balanserer" if abs(total) <= 0.01 else f"FEIL – sum = {total:.2f} (må være 0)"

    return (
        f"Forhåndsvisning GL-bilag\n"
        f"{'='*50}\n"
        f"Periode:        {periode_int}\n"
        f"Bokføringsdato: {dato.strftime('%d.%m.%Y')}\n"
        f"Batch ID:       {batch_id} (neste ledige)\n"
        f"Antall linjer:  {len(linjer)}\n"
        f"\nPosteringslinjer:\n" + "\n".join(lines_out) +
        f"\n\nSaldo: {total:.2f}  →  {status}"
    )


@mcp.tool()
def create_gl_bilag(
    periode: str,
    bokføringsdato: str,
    linjer_json: str,
    bilagsnavn: str = "",
) -> str:
    """
    Generer Excelerator GL-bilag som .xlsx på Skrivebordet, klar for import via Citrix.

    Args:
        periode: Periode som YYYYMM eller månedsnavn (f.eks. "202603" eller "mars 2026")
        bokføringsdato: Bokføringsdato som DD.MM.YYYY eller YYYY-MM-DD
        linjer_json: JSON-array med posteringslinjer. Hvert objekt må ha:
            account (int), amount (float), description (str).
            Valgfritt: dim_1 (koststed), dim_2 (prosjekt), dim_3 (ressursnr/lev.nr),
                       dim_4, dim_6, tax_code.
            Eksempel: [{"account": 29300, "amount": -9789.0, "description": "Kryssing"},
                       {"account": 29300, "amount":  9789.0, "description": "Kryssing"}]
        bilagsnavn: Valgfritt filnavn uten .xlsx. Autogenereres hvis tomt.
    """
    try:
        periode_int = _parse_periode(periode)
        dato = _parse_dato(bokføringsdato)
        linjer = json.loads(linjer_json)
    except Exception as e:
        return f"Feil i input: {e}"

    try:
        filsti = _generer_fil(
            periode=periode_int,
            bokføringsdato=dato,
            linjer=linjer,
            bilagsnavn=bilagsnavn if bilagsnavn.strip() else None,
        )
    except ValueError as e:
        return f"Valideringsfeil: {e}"
    except Exception as e:
        return f"Kunne ikke generere fil: {e}"

    total = sum(l["amount"] for l in linjer)
    return (
        f"GL-bilag generert:\n"
        f"  Fil:    {filsti}\n"
        f"  Sum:    {total:.2f} (skal være 0.00)\n"
        f"  Linjer: {len(linjer)}\n\n"
        f"Neste steg: Åpne filen i Excel inne i Citrix og importer via Excelerator."
    )


@mcp.tool()
def list_gl_bilag() -> str:
    """List alle Excelerator GL-bilag (.xlsx med _control-ark) på Skrivebordet."""
    try:
        import openpyxl
    except ImportError:
        return "openpyxl er ikke installert. Kjør: uv run --with openpyxl python server.py"

    funn = []
    for fil in sorted(DESKTOP.glob("*.xlsx")):
        if fil.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(fil, data_only=True)
        except Exception:
            continue
        if "_control" not in wb.sheetnames:
            continue
        ws = wb["_control"]
        batch_id = periode = dato = None
        for row in ws.iter_rows(values_only=True):
            if row[0] == "setdefault":
                if row[1] == "batch_id":
                    batch_id = row[2]
                elif row[1] == "period":
                    periode = row[2]
                elif row[1] == "voucher_date":
                    dato = row[2]
        ws_p = wb["Postering til UBW"] if "Postering til UBW" in wb.sheetnames else None
        antall_linjer = 0
        if ws_p:
            for rad in ws_p.iter_rows(min_row=10, values_only=True):
                if rad[0] == "update_data":
                    antall_linjer += 1
        dato_str = dato.strftime("%d.%m.%Y") if isinstance(dato, datetime) else str(dato)
        funn.append(f"  {fil.name}\n    batch_id={batch_id}  periode={periode}  dato={dato_str}  linjer={antall_linjer}")

    if not funn:
        return "Ingen GL-bilag funnet på Skrivebordet."
    return f"GL-bilag på Skrivebordet ({len(funn)} stk):\n\n" + "\n\n".join(funn)


if __name__ == "__main__":
    mcp.run()
